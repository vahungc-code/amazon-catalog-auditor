"""
CLR Parser Module
Extracts and normalizes data from Amazon Category Listing Reports
"""

import openpyxl
from typing import Optional, Dict, List
from dataclasses import dataclass


@dataclass
class Listing:
    """Normalized listing data"""
    row_number: int
    sku: str
    product_type: str
    item_type: str
    title: str
    brand: str
    parentage: str
    parent_sku: str
    status: str
    bullet_points: List[str]
    all_fields: Dict[str, any]  # Full row data


class CLRParser:
    """Parse Amazon Category Listing Reports"""

    # Standard CLR row structure
    ROW_SETTINGS = 1
    ROW_INSTRUCTIONS = 2
    ROW_GROUP_HEADERS = 3
    ROW_COL_HEADERS = 4
    ROW_FIELD_IDS = 5
    ROW_EXAMPLE = 6
    ROW_DATA_START = 7

    # Amazon-controlled or auto-populated fields: exclude from completeness
    # score and missing-attribute checks.
    # - listing_status: controlled by Amazon (Active/Inactive/Removed)
    # - title: controlled by Amazon (the display title)
    # - contribution_sku: the SKU itself — always present if the row exists
    # - record_action: internal Amazon action flag, not a seller attribute
    AMAZON_CONTROLLED_FIELDS = {'listing_status', 'title', 'contribution_sku', 'record_action'}

    # Known translations of "Active" status across CLR languages
    ACTIVE_STATUS_VALUES = {
        'active', 'attiva', 'attivo', 'aktiv', 'activo', 'activa',
        'actif', 'active', 'ativo', 'ativa', 'アクティブ', '有效', 'نشط',
    }

    # Mapping from Row 5 field ID prefixes (always English) to logical column names.
    # Used to resolve columns regardless of the Row 4 display language.
    FIELD_ID_TO_COLUMN = {
        'contribution_sku':        'SKU',
        '::listing_status':        'Status',
        '::title':                 'Title',
        'product_type':            'Product Type',
        'item_type_keyword':       'Item Type Keyword',
        'brand':                   'Brand',
        'parentage_level':         'Parentage',
        'child_parent_sku_relationship': 'Parent SKU',
        '::record_action':         'Record Action',
    }
    
    # Known translations of the "Template" sheet name across Amazon CLR languages
    TEMPLATE_SHEET_NAMES = [
        'Template',           # English
        'Modello',            # Italian
        'Vorlage',            # German
        'Plantilla',          # Spanish
        'Modèle',             # French
        'Modelo',             # Portuguese
        'テンプレート',          # Japanese
        '模板',                # Chinese
        'قالب',               # Arabic
    ]

    # Known translations of the "Data Definitions" sheet name
    DATA_DEFINITIONS_SHEET_NAMES = [
        'Data Definitions',
        'Definizioni dati',           # Italian
        'Datendefinitionen',          # German
        'Definiciones de datos',      # Spanish
        'Définitions de données',     # French
        'Definições de dados',        # Portuguese
        'データ定義',                   # Japanese
        '数据定义',                     # Chinese
        'تعريفات البيانات',             # Arabic
    ]

    def __init__(self, clr_file_path: str):
        """Load and parse CLR file"""
        self.file_path = clr_file_path
        self.workbook = openpyxl.load_workbook(clr_file_path, data_only=True, read_only=True)

        # Load sheets (language-agnostic lookup)
        self.template_sheet = self._find_template_sheet()
        
        # Parse structure
        # Row 4 = display headers (translated), Row 5 = field IDs (always English)
        self.headers = self._extract_headers()       # Row 4 display names → col index
        self.field_ids = self._extract_field_ids()    # Row 5 field IDs → col index
        # Resolved column map: logical English name → col index (uses Row 5 first, Row 4 fallback)
        self.columns = self._build_column_map()
        # Display names: col index → Row 4 display name (for UI, may be translated)
        self.display_names = {idx: name for name, idx in self.headers.items()}
        self.field_definitions = self._extract_field_definitions()
        
    def _find_template_sheet(self):
        """Find the Template sheet regardless of language.
        Tries known translations first, then falls back to the first sheet
        that isn't a Data Definitions sheet."""
        for name in self.TEMPLATE_SHEET_NAMES:
            if name in self.workbook.sheetnames:
                return self.workbook[name]

        # Fallback: pick the first sheet that isn't a known Data Definitions name
        dd_names = set(n.lower() for n in self.DATA_DEFINITIONS_SHEET_NAMES)
        for name in self.workbook.sheetnames:
            if name.lower() not in dd_names:
                return self.workbook[name]

        raise KeyError(
            f"Could not find a Template sheet. Available sheets: {self.workbook.sheetnames}"
        )

    def _find_data_definitions_sheet(self):
        """Find the Data Definitions sheet regardless of language.
        Returns None if not found."""
        for name in self.DATA_DEFINITIONS_SHEET_NAMES:
            if name in self.workbook.sheetnames:
                return self.workbook[name]

        # Fallback: pick a sheet that isn't the template sheet
        template_name = self.template_sheet.title
        for name in self.workbook.sheetnames:
            if name != template_name:
                return self.workbook[name]

        return None

    def _extract_headers(self) -> Dict[str, int]:
        """Extract column headers and their positions"""
        headers = {}
        header_row = self.template_sheet[self.ROW_COL_HEADERS]
        
        for idx, cell in enumerate(header_row, start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx
        
        return headers
    
    def _extract_field_ids(self) -> Dict[str, int]:
        """Extract field IDs from Row 5 and map them to column indices.
        These IDs match the Data Definitions 'Field Name' column."""
        field_ids = {}
        field_id_row = self.template_sheet[self.ROW_FIELD_IDS]

        for idx, cell in enumerate(field_id_row, start=1):
            if cell.value:
                field_ids[str(cell.value).strip()] = idx

        return field_ids

    @staticmethod
    def _field_id_base(field_id: str) -> str:
        """Extract the base field name from a Row 5 field ID.
        e.g. 'parentage_level[marketplace_id=APJ6JRA9NG5V4]#1.value' → 'parentage_level'
             'contribution_sku#1.value' → 'contribution_sku'
             '::listing_status' → '::listing_status'
        """
        # Strip [marketplace_id=...] brackets, then #N.value suffixes
        base = field_id.split('[')[0].split('#')[0].split('.')[0]
        return base

    def _build_column_map(self) -> Dict[str, int]:
        """Build a logical column map using Row 5 field IDs (always English).
        Maps logical English names (e.g. 'SKU', 'Status') to column indices.
        Falls back to Row 4 headers for anything not matched via Row 5."""
        columns = {}

        # First, resolve from Row 5 field IDs
        for field_id, col_idx in self.field_ids.items():
            base = self._field_id_base(field_id)
            if base in self.FIELD_ID_TO_COLUMN:
                logical_name = self.FIELD_ID_TO_COLUMN[base]
                # Keep the first match (lowest column index) for each logical name
                if logical_name not in columns:
                    columns[logical_name] = col_idx

        # Also resolve bullet points from Row 5 field IDs
        for field_id, col_idx in self.field_ids.items():
            base = self._field_id_base(field_id)
            if base == 'bullet_point':
                # Extract the number from e.g. bullet_point#1.value
                try:
                    num = int(field_id.split('#')[1].split('.')[0])
                    columns[f'Bullet Point {num}'] = col_idx
                except (IndexError, ValueError):
                    pass

        # Fall back to Row 4 headers for anything still missing
        for header_name, col_idx in self.headers.items():
            if header_name not in columns:
                columns[header_name] = col_idx

        return columns

    # Multilingual translations of Data Definitions column headers
    DD_FIELD_NAME_HEADERS = {
        'field name', 'nome del campo', 'feldname', 'nombre del campo',
        'nom du champ', 'nome do campo', 'フィールド名', '字段名称', 'اسم الحقل',
    }
    DD_REQUIRED_HEADERS = {
        'required?', 'obbligatorio?', 'erforderlich?', '¿obligatorio?',
        'obligatoire ?', 'obrigatório?', '必須?', '必填?', 'مطلوب؟',
    }
    # Values that mean "required" across languages
    REQUIRED_VALUES = {
        'required', 'obbligatorio', 'erforderlich', 'obligatorio',
        'obligatoire', 'obrigatório', '必須', '必填', 'مطلوب',
    }

    def _extract_field_definitions(self) -> Dict[str, Dict]:
        """Extract field definitions from Data Definitions sheet (any language)"""
        definitions = {}

        dd_sheet = self._find_data_definitions_sheet()
        if dd_sheet is None:
            return definitions

        # Find header row by looking for any known translation of "Field Name"
        header_row_idx = None
        for i in range(1, 6):
            row = dd_sheet[i]
            for cell in row:
                if cell.value and str(cell.value).strip().lower() in self.DD_FIELD_NAME_HEADERS:
                    header_row_idx = i
                    break
            if header_row_idx:
                break

        if not header_row_idx:
            return definitions

        # Extract headers and resolve "Field Name" and "Required?" columns
        dd_headers = {}
        for idx, cell in enumerate(dd_sheet[header_row_idx], start=1):
            if cell.value:
                dd_headers[str(cell.value).strip().lower()] = idx

        # Find the field name column (any language)
        field_name_idx = None
        for key in self.DD_FIELD_NAME_HEADERS:
            if key in dd_headers:
                field_name_idx = dd_headers[key]
                break

        # Find the required column (any language)
        required_idx = None
        for key in self.DD_REQUIRED_HEADERS:
            if key in dd_headers:
                required_idx = dd_headers[key]
                break

        if not field_name_idx:
            return definitions

        # Extract definitions
        for row in dd_sheet.iter_rows(min_row=header_row_idx + 1):
            field_name = row[field_name_idx - 1].value
            if not field_name:
                continue

            required = row[required_idx - 1].value if required_idx else None
            required_lower = str(required).strip().lower() if required else ''

            definitions[str(field_name).strip()] = {
                'required': 'required' if required_lower in self.REQUIRED_VALUES else required_lower,
                'field_name': str(field_name).strip()
            }

        return definitions
    
    def _is_amazon_controlled(self, field_name: str) -> bool:
        """Check if a field is Amazon-controlled (not editable by sellers).
        Handles various naming formats:
          '::listing_status'  → listing_status
          'contribution_sku#1.value' → contribution_sku
          'title' → title
        """
        normalized = field_name.strip().lstrip(':').lower()
        # Strip suffixes like #1.value, #2.value etc.
        base_name = normalized.split('#')[0].split('.')[0]
        return normalized in self.AMAZON_CONTROLLED_FIELDS or base_name in self.AMAZON_CONTROLLED_FIELDS

    def get_required_fields(self) -> List[str]:
        """Get list of required field names (excluding Amazon-controlled fields)"""
        return [
            field_name
            for field_name, definition in self.field_definitions.items()
            if definition['required'] == 'required'
            and not self._is_amazon_controlled(field_name)
        ]

    def get_conditional_fields(self) -> List[str]:
        """Get list of conditionally required field names (excluding Amazon-controlled fields)"""
        return [
            field_name
            for field_name, definition in self.field_definitions.items()
            if 'conditional' in definition['required'].lower()
            and not self._is_amazon_controlled(field_name)
        ]
    
    def get_listings(self, skip_parents: bool = True, skip_examples: bool = True,
                     skip_fbm_duplicates: bool = True, active_only: bool = True) -> List[Listing]:
        """
        Extract all listings from CLR

        Args:
            skip_parents: Skip parent SKUs (variations)
            skip_examples: Skip example/dummy rows
            skip_fbm_duplicates: Skip FBM duplicates when FBA version exists
            active_only: Only include listings with "Active" status

        Returns:
            List of normalized Listing objects
        """
        listings = []
        
        # Get column indices via resolved column map (Row 5 field IDs, always English)
        col_sku = self.columns.get('SKU', 3)
        col_status = self.columns.get('Status', 1)
        col_title = self.columns.get('Title', 2)
        col_product_type = self.columns.get('Product Type', 4)
        col_item_type = self.columns.get('Item Type Keyword', 13)
        col_brand = self.columns.get('Brand', 10)
        col_parentage = self.columns.get('Parentage', 6)
        col_parent_sku = self.columns.get('Parent SKU', 7)

        # Bullet point columns
        bullet_cols = []
        for i in range(1, 6):
            col_name = f'Bullet Point {i}'
            if col_name in self.columns:
                bullet_cols.append(self.columns[col_name])
        
        # Iterate through data rows
        for row_idx, row in enumerate(self.template_sheet.iter_rows(min_row=self.ROW_DATA_START), 
                                      start=self.ROW_DATA_START):
            # Extract basic fields
            sku = self._get_cell_value(row, col_sku)
            
            if not sku:
                continue
            
            # Skip examples
            if skip_examples and sku.upper() in ['ABC123', 'EXAMPLE', 'TEST']:
                continue
            
            status = self._get_cell_value(row, col_status)
            parentage = self._get_cell_value(row, col_parentage)

            # Skip parents if requested
            if skip_parents and parentage and 'parent' in parentage.lower():
                continue

            # Skip non-active listings (Removed, Inactive, blank, etc.)
            if active_only:
                if not status or status.strip().lower() not in self.ACTIVE_STATUS_VALUES:
                    continue

            # Extract all fields — index by:
            # 1. Resolved logical English names from self.columns
            # 2. Row 4 display headers (may be translated)
            # 3. Row 5 field IDs (always English, e.g. "contribution_sku#1.value")
            all_fields = {}
            for logical_name, col_idx in self.columns.items():
                all_fields[logical_name] = self._get_cell_value(row, col_idx)
            for header_name, col_idx in self.headers.items():
                if header_name not in all_fields:
                    all_fields[header_name] = self._get_cell_value(row, col_idx)
            for field_id, col_idx in self.field_ids.items():
                if field_id not in all_fields:
                    all_fields[field_id] = self._get_cell_value(row, col_idx)
            
            # Extract bullet points
            bullets = []
            for bullet_col in bullet_cols:
                bullet_text = self._get_cell_value(row, bullet_col)
                bullets.append(bullet_text if bullet_text else "")
            
            # Create listing object
            listing = Listing(
                row_number=row_idx,
                sku=sku,
                product_type=self._get_cell_value(row, col_product_type) or "",
                item_type=self._get_cell_value(row, col_item_type) or "",
                title=self._get_cell_value(row, col_title) or "",
                brand=self._get_cell_value(row, col_brand) or "",
                parentage=parentage or "",
                parent_sku=self._get_cell_value(row, col_parent_sku) or "",
                status=status or "",
                bullet_points=bullets,
                all_fields=all_fields
            )
            
            listings.append(listing)
        
        # Filter FBM/MFN duplicates (keep FBA versions)
        if skip_fbm_duplicates:
            listings = self._filter_fbm_duplicates(listings)
        
        return listings
    
    def _filter_fbm_duplicates(self, listings: List[Listing]) -> List[Listing]:
        """
        Filter out FBM/MFN duplicates of FBA listings.
        When multiple SKUs have the same item name (title), keep the FBA version.
        """
        seen_names = {}
        filtered = []
        skipped_count = 0
        
        for listing in listings:
            # Use title as the unique identifier (item name)
            item_name = listing.title.strip() if listing.title else ""
            
            if not item_name:
                # No title, can't detect duplicates - keep it
                filtered.append(listing)
                continue
            
            if item_name in seen_names:
                # Duplicate found
                existing_sku = seen_names[item_name]
                
                # If this one is FBA, replace the existing one
                if "_FBA_" in listing.sku.upper() or "FBA" in listing.sku.upper():
                    # Remove the old one, add this FBA version
                    filtered = [l for l in filtered if l.sku != existing_sku]
                    filtered.append(listing)
                    seen_names[item_name] = listing.sku
                else:
                    # This is FBM/MFN, skip it
                    skipped_count += 1
                    continue
            else:
                # First time seeing this item name
                seen_names[item_name] = listing.sku
                filtered.append(listing)
        
        if skipped_count > 0:
            print(f"Skipped {skipped_count} FBM/MFN duplicates (keeping FBA versions)")
        
        return filtered
    
    def get_product_types(self) -> List[str]:
        """Get unique product types in catalog"""
        product_types = set()
        listings = self.get_listings()
        
        for listing in listings:
            if listing.product_type:
                product_types.add(listing.product_type)
        
        return sorted(list(product_types))
    
    def _get_cell_value(self, row, col_idx: int) -> Optional[str]:
        """Safely get cell value from row"""
        try:
            if col_idx <= 0 or col_idx > len(row):
                return None
            
            cell = row[col_idx - 1]
            if cell.value is None:
                return None
            
            return str(cell.value).strip()
        except (IndexError, AttributeError):
            return None
    
    def __del__(self):
        """Clean up workbook"""
        if hasattr(self, 'workbook'):
            self.workbook.close()
